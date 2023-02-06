import os

from celery.result import AsyncResult
from openpyxl.workbook import Workbook

from app.core.celery_app import celery
from app.core.config import ExcelStyle, settings


@celery.task
def generate_menu_xlsx(menus_data: list[dict]) -> AsyncResult:
    """Generate menu report in Excel (xlsx) file format."""

    task_id = celery.current_task.request.id
    filename = f"{task_id}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Меню"
    row = 1

    column_dimensions = {"A": 5, "B": 15, "C": 35, "D": 35, "E": 70, "F": 10}
    for col, value in column_dimensions.items():
        ws.column_dimensions[col].width = value

    for menu_num in range(len(menus_data)):
        for col_number in range(1, len(column_dimensions) + 1):
            ws.cell(row, col_number).style = ExcelStyle.Header
            ws.cell(row, col_number).fill = ExcelStyle.GreyFill
        ws.cell(row, 1).value = menu_num + 1
        ws.cell(row, 2).value = menus_data[menu_num]["title"]
        ws.cell(row, 3).value = menus_data[menu_num]["description"]
        submenus_data = menus_data[menu_num]["submenus"]
        row += 1
        for submenu_num in range(len(submenus_data)):
            for col_number in range(1, len(column_dimensions) + 1):
                ws.cell(row, col_number).style = ExcelStyle.BaseBold
            ws.cell(row, 2).value = submenu_num + 1
            ws.cell(row, 3).value = submenus_data[submenu_num]["title"]
            ws.cell(row, 4).value = submenus_data[submenu_num]["description"]
            dishes_data = submenus_data[submenu_num]["dishes"]
            row += 1
            for dish_num in range(len(dishes_data)):
                for col_number in range(1, len(column_dimensions) + 1):
                    ws.cell(row, col_number).style = ExcelStyle.Base
                ws.cell(row, 3).value = dish_num + 1
                ws.cell(row, 4).value = dishes_data[dish_num]["title"]
                ws.cell(row, 5).value = dishes_data[dish_num]["description"]
                ws.cell(row, 6).value = dishes_data[dish_num]["price"]
                row += 1

    wb.save(os.path.join(settings.FILES_DIR, filename))
    return task_id


def get_task_info(task_id: str) -> dict:
    """Returns task info for the given task_id."""

    task_result = AsyncResult(task_id)
    result = {
        "task_id": task_result.id.__str__(),
        "task_status": task_result.status,
        "task_result": task_result.result,
    }
    return result
